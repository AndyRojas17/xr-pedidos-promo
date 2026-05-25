import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils import render_header, render_footer

st.set_page_config(
    page_title="En Tránsito — XR Panel",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

render_header("En Tránsito — Consulta de pedidos y lista promo")

# ── CARGA DE DATOS ────────────────────────────────────────────────────────────
DATA_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "en_transito.csv")

@st.cache_data
def cargar_datos():
    df = pd.read_csv(DATA_PATH, dtype=str).fillna('')
    df['Cantidad']  = pd.to_numeric(df['Cantidad'],  errors='coerce').fillna(0)
    df['Precio']    = pd.to_numeric(df['Precio'],    errors='coerce')
    df['Descuento'] = pd.to_numeric(df['Descuento'], errors='coerce')
    df['Total']     = pd.to_numeric(df['Total'],     errors='coerce')
    return df

df = cargar_datos()
df_transito = df[df['Estado'] == 'En Transito']

# variables para uso en descarga excel
total_items     = len(df_transito)
total_inversion = df_transito['Total'].sum()

st.markdown('<div style="margin-top:8px"></div>', unsafe_allow_html=True)

# ── BUSCADOR ─────────────────────────────────────────────────────────────────
st.markdown("""
<p style="color:#888;font-size:0.85rem;margin-bottom:12px">
    Escribe el código o nombre del repuesto para saber si fue pedido, su precio y descuento.
</p>
""", unsafe_allow_html=True)

query = st.text_input("", placeholder="Ej: 12391KRM840   /   pastilla freno   /   CB190R",
                      label_visibility="collapsed")

st.markdown('<div style="margin-top:8px"></div>', unsafe_allow_html=True)

# ── RESULTADOS ────────────────────────────────────────────────────────────────
if query.strip():
    terminos = [t.lower().strip() for t in query.split() if t.strip()]

    resultados = []
    for _, row in df.iterrows():
        texto = ' '.join([
            str(row['Codigo']).lower(),
            str(row['Descripcion']).lower(),
            str(row['Modelo']).lower(),
        ])
        matches = sum(1 for t in terminos if t in texto)
        if matches > 0:
            resultados.append((matches, row))

    resultados.sort(key=lambda x: (-x[0], x[1]['Estado'] != 'En Transito'))

    if not resultados:
        st.warning(f"No se encontró **\"{query}\"** en la lista promocional.")
    else:
        st.markdown(f"**{len(resultados)} resultado(s)** para *\"{query}\"*")
        st.markdown('<div style="margin-top:8px"></div>', unsafe_allow_html=True)

        for _, row in resultados:
            en_transito = row['Estado'] == 'En Transito'

            precio_str    = f"S/. {row['Precio']:.2f}"   if pd.notna(row['Precio'])    else '—'
            desc_str      = f"{row['Descuento']*100:.1f}%" if pd.notna(row['Descuento']) else '—'
            cantidad_str  = f"{int(row['Cantidad'])} uds"

            if en_transito:
                borde_color  = '#2E7D32'
                estado_bg    = '#E8F5E9'
                estado_fg    = '#2E7D32'
                estado_txt   = '✅ En Tránsito'
                detalle_html = (f'<span style="font-size:0.8rem;color:#666">📦 <b style="color:#1A1A1A">{cantidad_str}</b></span>'
                                f'&nbsp;&nbsp;<span style="font-size:0.8rem;color:#666">💲 Precio: <b>{precio_str}</b></span>'
                                f'&nbsp;&nbsp;<span style="font-size:0.8rem;color:#666">🏷️ Desc.: <b style="color:#CC0000">{desc_str}</b></span>')
            else:
                borde_color  = '#AAAAAA'
                estado_bg    = '#F5F5F5'
                estado_fg    = '#888888'
                estado_txt   = '⚠️ No solicitado'
                detalle_html = (f'<span style="font-size:0.8rem;color:#666">💲 Precio lista: <b>{precio_str}</b></span>'
                                f'&nbsp;&nbsp;<span style="font-size:0.8rem;color:#666">🏷️ Desc. promo: <b style="color:#CC0000">{desc_str}</b></span>'
                                f'&nbsp;&nbsp;<span style="font-size:0.8rem;color:#888;font-style:italic">Este repuesto no fue incluido en el pedido.</span>')

            modelo_txt = row['Modelo'] if row['Modelo'] else '—'
            card = (
                f'<div style="background:#FAFAFA;border:1px solid #E5E5E5;border-left:4px solid {borde_color};border-radius:8px;padding:14px 20px;margin-bottom:8px">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;margin-bottom:6px">'
                f'<div style="flex:1;min-width:200px">'
                f'<span style="font-family:monospace;background:#F0F0F0;color:#333;padding:2px 8px;border-radius:4px;font-size:0.85rem;font-weight:700">{row["Codigo"]}</span>'
                f'<span style="font-size:0.95rem;font-weight:700;color:#1A1A1A;margin-left:10px">{row["Descripcion"]}</span>'
                f'</div>'
                f'<span style="background:{estado_bg};color:{estado_fg};padding:3px 12px;border-radius:10px;font-size:0.78rem;font-weight:700;white-space:nowrap">{estado_txt}</span>'
                f'</div>'
                f'<div style="margin-bottom:8px"><span style="font-size:0.78rem;color:#888">🏍️ {modelo_txt}</span></div>'
                f'<div style="display:flex;flex-wrap:wrap;gap:12px;align-items:center">{detalle_html}</div>'
                f'</div>'
            )
            st.markdown(card, unsafe_allow_html=True)

else:
    # Estado inicial — mostrar ítems en tránsito
    st.markdown("### Ítems actualmente en tránsito")
    st.markdown('<div style="margin-top:4px"></div>', unsafe_allow_html=True)

    for _, row in df_transito.iterrows():
        precio_str = f"S/. {row['Precio']:.2f}" if pd.notna(row['Precio']) else '—'
        desc_str   = f"{row['Descuento']*100:.1f}%" if pd.notna(row['Descuento']) else '—'
        card = (
            f'<div style="background:#FAFAFA;border:1px solid #E5E5E5;border-left:4px solid #2E7D32;border-radius:8px;padding:12px 20px;margin-bottom:6px">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px">'
            f'<div style="flex:1;min-width:200px">'
            f'<span style="font-family:monospace;background:#F0F0F0;color:#333;padding:2px 8px;border-radius:4px;font-size:0.82rem;font-weight:700">{row["Codigo"]}</span>'
            f'<span style="font-size:0.9rem;font-weight:700;color:#1A1A1A;margin-left:10px">{row["Descripcion"]}</span>'
            f'</div>'
            f'<div style="display:flex;gap:16px;flex-wrap:wrap;align-items:center">'
            f'<span style="font-size:0.8rem;color:#666">📦 <b>{int(row["Cantidad"])} uds</b></span>'
            f'<span style="font-size:0.8rem;color:#666">💲 <b>{precio_str}</b></span>'
            f'<span style="font-size:0.8rem;color:#666">🏷️ <b style="color:#CC0000">{desc_str}</b></span>'
            f'</div></div></div>'
        )
        st.markdown(card, unsafe_allow_html=True)

# ── DESCARGA ──────────────────────────────────────────────────────────────────
st.markdown('<div style="margin-top:32px"></div>', unsafe_allow_html=True)
st.markdown('<hr style="border:none;border-top:1px solid #EEE;margin-bottom:20px">', unsafe_allow_html=True)

def generar_excel_transito(df_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'En Tránsito'

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'XR MOTO STORE — PEDIDOS EN TRÁNSITO + LISTA PROMO'
    c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='CC0000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers = ['Código', 'Descripción', 'Modelo', 'Precio (S/.)', 'Desc. %', 'Cantidad', 'Total (S/.)', 'Estado']
    widths  = [18, 40, 30, 14, 10, 12, 14, 18]
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='404040')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    fill_transito = PatternFill('solid', start_color='E8F5E9')
    fill_no_solic = PatternFill('solid', start_color='F5F5F5')

    for ri, (_, row) in enumerate(df_data.iterrows()):
        er = ri + 3
        precio   = row['Precio']    if pd.notna(row['Precio'])    else ''
        desc     = row['Descuento'] if pd.notna(row['Descuento']) else ''
        cantidad = int(row['Cantidad']) if row['Cantidad'] > 0 else ''
        total    = row['Total']     if pd.notna(row['Total'])     else ''
        vals = [row['Codigo'], row['Descripcion'], row['Modelo'],
                precio, desc, cantidad, total, row['Estado']]
        fill = fill_transito if row['Estado'] == 'En Transito' else fill_no_solic

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.border = border
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='center')
            cell.fill = fill
            if ci == 1:
                cell.font = Font(name='Arial', bold=True, size=9)
            if ci == 4 and val != '':
                cell.number_format = '#,##0.00'
            if ci == 5 and val != '':
                cell.number_format = '0.0%'
            if ci == 7 and val != '':
                cell.number_format = '#,##0.00'
        ws.row_dimensions[er].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

col_dl, col_info = st.columns([1, 2])
with col_dl:
    clave = st.text_input("🔒 Clave de administrador", type="password",
                          placeholder="Ingresa la clave para descargar")
    if clave == "xr3010":
        excel_data = generar_excel_transito(df)
        st.download_button(
            label="⬇️  Descargar lista completa (Excel)",
            data=excel_data,
            file_name="XR_En_Transito.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    elif clave:
        st.markdown('<p style="color:#CC0000;font-size:0.82rem;margin-top:4px">Clave incorrecta.</p>',
                    unsafe_allow_html=True)
with col_info:
    st.markdown(f"""
    <p style="color:#888;font-size:0.85rem;margin-top:8px">
        Base completa: <b>{len(df)} ítems</b> —
        <b>{total_items} pedidos</b> en tránsito y
        <b>{len(df)-total_items} no solicitados</b>.
        Solo disponible para administradores.
    </p>
    """, unsafe_allow_html=True)

# ── CARGAR NUEVA LISTA ────────────────────────────────────────────────────────
st.markdown('<div style="margin-top:12px"></div>', unsafe_allow_html=True)

def detectar_hoja_y_skip(bytes_data):
    """Detecta la hoja correcta y la fila de encabezados en cualquier Excel de lista promo."""
    wb_tmp = openpyxl.load_workbook(io.BytesIO(bytes_data), read_only=True, data_only=True)
    for nombre_hoja in wb_tmp.sheetnames:
        ws_tmp = wb_tmp[nombre_hoja]
        for i, row in enumerate(ws_tmp.iter_rows(max_row=20, values_only=True)):
            fila = ' '.join(str(c) for c in row if c is not None)
            # Busca la fila que tenga columna de cantidad pedida Y columna de código
            tiene_pedido = 'PEDIDO' in fila.upper() or 'CANTIDAD' in fila.upper() or 'QTY' in fila.upper()
            tiene_codigo = any(p in fila for p in ['aterial', 'ODIGO', 'Codigo', 'codigo', 'CODE'])
            if tiene_pedido and tiene_codigo:
                wb_tmp.close()
                return nombre_hoja, i
    wb_tmp.close()
    return wb_tmp.sheetnames[0], 5  # fallback

def detectar_hoja_paty(bytes_data):
    """Detecta la hoja con sugerencias Paty (busca columna SUGERENCIA o PATY)."""
    wb_tmp = openpyxl.load_workbook(io.BytesIO(bytes_data), read_only=True, data_only=True)
    for nombre_hoja in wb_tmp.sheetnames:
        ws_tmp = wb_tmp[nombre_hoja]
        for row in ws_tmp.iter_rows(max_row=10, values_only=True):
            fila = ' '.join(str(c) for c in row if c is not None).upper()
            if 'SUGERENCIA' in fila or 'PATY' in fila:
                wb_tmp.close()
                return nombre_hoja
    hoja_fallback = wb_tmp.sheetnames[0]
    wb_tmp.close()
    return hoja_fallback  # primera hoja por defecto

def procesar_lista(promo_bytes, paty_bytes=None):
    """Procesa los Excel y devuelve el DataFrame listo para guardar."""
    hoja, skip = detectar_hoja_y_skip(promo_bytes)

    # ─ Lista Promocional ─
    promo = pd.read_excel(io.BytesIO(promo_bytes), sheet_name=hoja, skiprows=skip, dtype=str)
    promo.columns = promo.columns.str.strip()

    cols_up = [c.upper().strip() for c in promo.columns]

    def buscar_col(palabras_clave, excluir=None):
        for c, cu in zip(promo.columns, cols_up):
            if any(p.upper() in cu for p in palabras_clave):
                if excluir and any(e.upper() in cu for e in excluir):
                    continue
                return c
        return None

    codigo_col  = buscar_col(['MATERIAL', 'CODIGO', 'CÓDIGO', 'CODE', 'COD.', 'N.MAT'])
    pedido_col  = buscar_col(['PEDIDO', 'CANTIDAD', 'QTY', 'ORDER'], excluir=['US$', 'USD', 'DOLAR', '$'])
    desc_col    = buscar_col(['DESCRIPCION', 'DESCRIPCIÓN', 'DESCRIPTION', 'DETALLE'])
    modelo_col  = buscar_col(['MODELO', 'MODEL', 'APLICACION', 'APLICACIÓN', 'APLICA'])
    precio_col  = buscar_col(['OFERTA S', 'PRECIO S/', 'PRICE S/', 'PVP S/', 'SOLES', 'S/.'])
    dpct_col    = buscar_col(['DESC. %', 'DESCUENTO', 'DISCOUNT', 'DESC%', '% DESC'])

    if not codigo_col or not pedido_col:
        raise ValueError(f"No se encontraron las columnas de código o pedido. Columnas disponibles: {promo.columns.tolist()}")

    rename_map = {codigo_col: 'Codigo', pedido_col: 'Pedido'}
    if desc_col:   rename_map[desc_col]   = 'Descripcion'
    if modelo_col: rename_map[modelo_col] = 'Modelo'
    if precio_col: rename_map[precio_col] = 'Precio'
    if dpct_col:   rename_map[dpct_col]   = 'Descuento'

    promo = promo.rename(columns=rename_map)

    for col in ['Descripcion', 'Modelo', 'Precio', 'Descuento']:
        if col not in promo.columns:
            promo[col] = ''

    promo['Codigo']     = promo['Codigo'].astype(str).str.strip()
    promo['Pedido']     = pd.to_numeric(promo['Pedido'],     errors='coerce').fillna(0)
    promo['Precio']     = pd.to_numeric(promo['Precio'],     errors='coerce')
    promo['Descuento']  = pd.to_numeric(promo['Descuento'],  errors='coerce')
    promo['Descripcion']= promo['Descripcion'].fillna('').astype(str).str.strip()
    promo['Modelo']     = promo['Modelo'].fillna('').astype(str).str.strip()

    # Filtrar filas vacías o de encabezado repetido
    promo = promo[promo['Codigo'].str.len() > 3].copy()

    # ─ Sugeridos Paty ─
    if paty_bytes:
        hoja_paty = detectar_hoja_paty(paty_bytes)
        paty = pd.read_excel(io.BytesIO(paty_bytes), sheet_name=hoja_paty, dtype=str)
        paty.columns = paty.columns.str.strip()
        cols_paty_up = [c.upper() for c in paty.columns]

        paty_col = next((c for c, cu in zip(paty.columns, cols_paty_up)
                         if 'SUGERENCIA' in cu or 'PATY' in cu or 'CANTIDAD' in cu or 'QTY' in cu), None)
        cod_col  = next((c for c, cu in zip(paty.columns, cols_paty_up)
                         if any(p in cu for p in ['MATERIAL', 'CODIGO', 'CÓDIGO', 'COD', 'CODE'])), None)
        if not paty_col or not cod_col:
            raise ValueError(f"No se encontraron columnas en el archivo Paty. Columnas: {paty.columns.tolist()}")
        paty = paty.rename(columns={cod_col: 'Codigo', paty_col: 'Paty'})
        paty['Paty']   = pd.to_numeric(paty['Paty'], errors='coerce').fillna(0)
        paty['Codigo'] = paty['Codigo'].astype(str).str.strip()
        paty = paty[paty['Paty'] > 0][['Codigo', 'Paty']]

        promo = promo.merge(paty, on='Codigo', how='left')
        promo['Paty']       = promo['Paty'].fillna(0)
        promo['Fuente_Paty']= promo['Paty'].apply(lambda x: 'Si' if x > 0 else 'No')
        promo['Cantidad']   = promo['Pedido'] + promo['Paty']
        promo = promo.drop(columns=['Paty'])
    else:
        promo['Cantidad']    = promo['Pedido']
        promo['Fuente_Paty'] = 'No'

    promo['Estado'] = promo['Cantidad'].apply(lambda x: 'En Transito' if x > 0 else 'No solicitado')
    promo['Total']  = promo['Precio'] * promo['Cantidad']
    promo = promo.drop(columns=['Pedido'])

    # Ordenar: En Tránsito primero, luego por Cantidad desc
    promo['_ord'] = promo['Estado'].apply(lambda x: 0 if x == 'En Transito' else 1)
    promo = promo.sort_values(['_ord', 'Cantidad'], ascending=[True, False]).drop(columns=['_ord'])
    promo = promo.reset_index(drop=True)

    cols = ['Codigo', 'Descripcion', 'Modelo', 'Precio', 'Descuento', 'Cantidad', 'Total', 'Estado', 'Fuente_Paty']
    return promo[cols]

with st.expander("🔄  Cargar nueva lista de pedidos"):
    st.markdown('<p style="color:#888;font-size:0.85rem;margin-bottom:4px">Solo administradores. La lista actual se reemplaza al procesar.</p>', unsafe_allow_html=True)

    clave_up = st.text_input("🔒 Clave", type="password",
                             placeholder="Clave de administrador",
                             key="clave_upload")

    if clave_up == "xr3010":
        st.markdown('<div style="margin-top:8px"></div>', unsafe_allow_html=True)
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown('<p style="font-size:0.85rem;font-weight:700;margin-bottom:4px">📋 Lista Promocional Honda <span style="color:#CC0000">*</span></p>', unsafe_allow_html=True)
            promo_file = st.file_uploader("Lista Promo", type=['xlsx','xls'],
                                          label_visibility="collapsed", key="upload_promo")
        with col_f2:
            st.markdown('<p style="font-size:0.85rem;font-weight:700;margin-bottom:4px">📋 Sugeridos Paty <span style="color:#888;font-weight:400">(opcional)</span></p>', unsafe_allow_html=True)
            paty_file  = st.file_uploader("Sugeridos Paty", type=['xlsx','xls'],
                                          label_visibility="collapsed", key="upload_paty")

        st.markdown('<div style="margin-top:8px"></div>', unsafe_allow_html=True)
        if st.button("⚙️  Procesar y actualizar lista", use_container_width=False,
                     disabled=(promo_file is None)):
            with st.spinner("Procesando..."):
                try:
                    promo_bytes = promo_file.read()
                    paty_bytes  = paty_file.read() if paty_file else None
                    nuevo_df = procesar_lista(promo_bytes, paty_bytes)
                    nuevo_df.to_csv(DATA_PATH, index=False, encoding='utf-8')
                    st.cache_data.clear()
                    en_t = len(nuevo_df[nuevo_df['Estado'] == 'En Transito'])
                    st.success(f"✅ Lista actualizada: **{len(nuevo_df)} ítems** cargados — **{en_t} en tránsito**. Recarga la página para ver los cambios.")
                except Exception as e:
                    st.error(f"Error al procesar: {e}")

    elif clave_up:
        st.markdown('<p style="color:#CC0000;font-size:0.82rem;margin-top:4px">Clave incorrecta.</p>',
                    unsafe_allow_html=True)

render_footer()

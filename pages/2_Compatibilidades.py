import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils import render_header, render_footer

st.set_page_config(
    page_title="Compatibilidades — XR Panel",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

render_header("Compatibilidades — Buscador de repuestos y modelos")

# ── CARGA DE DATOS ────────────────────────────────────────────────────────────
DATA_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "repuestos.csv")

@st.cache_data
def cargar_datos():
    df = pd.read_csv(DATA_PATH, dtype=str).fillna('')
    df['Modelos_lista'] = df['Modelos'].apply(
        lambda x: [m.strip() for m in x.split('/') if m.strip()]
    )
    return df

df_rep = cargar_datos()

# ── BUSCADOR ──────────────────────────────────────────────────────────────────
st.markdown("""
<p style="color:#888;font-size:0.85rem;margin-bottom:16px">
    Escribe el modelo, nombre del repuesto o código. Puedes combinar términos:
    <code style="background:#F0F0F0;padding:1px 5px;border-radius:3px">XR190 freno</code> &nbsp;·&nbsp;
    <code style="background:#F0F0F0;padding:1px 5px;border-radius:3px">pastilla CB300</code> &nbsp;·&nbsp;
    <code style="background:#F0F0F0;padding:1px 5px;border-radius:3px">12391K70601</code>
</p>
""", unsafe_allow_html=True)

col_search, col_btn = st.columns([5, 1])
with col_search:
    query = st.text_input("", placeholder="Ej: XR190 freno   /   pastilla CB300   /   14401KVJ911",
                          label_visibility="collapsed")
with col_btn:
    buscar = st.button("🔍  Buscar", use_container_width=True)

st.markdown('<div style="margin-top:8px"></div>', unsafe_allow_html=True)

# ── LÓGICA DE BÚSQUEDA ────────────────────────────────────────────────────────
def buscar_repuestos(query_str, df):
    terminos = [t.lower().strip() for t in query_str.split() if t.strip()]
    if not terminos:
        return pd.DataFrame()

    resultados = []
    for _, row in df.iterrows():
        texto_busqueda = ' '.join([
            row['Codigo'].lower(),
            row['Descripcion'].lower(),
            row['Modelos'].lower(),
        ])
        matches = sum(1 for t in terminos if t in texto_busqueda)
        if matches > 0:
            resultados.append({
                '_matches': matches,
                'Código': row['Codigo'],
                'Descripción': row['Descripcion'],
                'Modelos compatibles': row['Modelos'],
                'Fuente': row['Fuente'],
                '_modelos_lista': row['Modelos_lista'],
            })

    if not resultados:
        return pd.DataFrame()

    df_res = pd.DataFrame(resultados)
    df_res = df_res.sort_values('_matches', ascending=False).reset_index(drop=True)
    df_res.index += 1
    return df_res

# ── RESULTADOS ────────────────────────────────────────────────────────────────
if buscar and query.strip():
    df_found = buscar_repuestos(query, df_rep)

    if df_found.empty:
        st.warning(f"No se encontraron resultados para **\"{query}\"**. Intenta con otros términos.")
    else:
        st.markdown(f"**{len(df_found)} resultado(s)** para *\"{query}\"*")
        st.markdown('<div style="margin-top:12px"></div>', unsafe_allow_html=True)

        for i, row in df_found.iterrows():
            modelos = row['_modelos_lista']
            modelos_html = ''.join([
                f'<span style="display:inline-block;background:#E8F0FE;color:#1A4FAE;'
                f'padding:3px 10px;border-radius:12px;font-size:0.8rem;font-weight:600;'
                f'margin:2px 4px 2px 0">{m}</span>'
                for m in modelos
            ]) if modelos else '<span style="color:#AAA;font-size:0.85rem">Sin info de modelos</span>'

            st.markdown(f"""
            <div style="background:#FAFAFA;border:1px solid #E5E5E5;border-left:4px solid #CC0000;
                        border-radius:8px;padding:16px 20px;margin-bottom:10px">
                <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px">
                    <div>
                        <span style="font-family:monospace;background:#F0F0F0;color:#333;
                                     padding:2px 8px;border-radius:4px;font-size:0.85rem;
                                     font-weight:700">{row['Código']}</span>
                        <span style="font-size:1rem;font-weight:700;color:#1A1A1A;
                                     margin-left:12px">{row['Descripción']}</span>
                    </div>
                    <span style="font-size:0.72rem;color:#AAAAAA;white-space:nowrap;
                                 margin-left:16px">{row['Fuente']}</span>
                </div>
                <div style="margin-top:6px">
                    <span style="font-size:0.78rem;color:#888;margin-right:8px">Modelos:</span>
                    {modelos_html}
                </div>
            </div>
            """, unsafe_allow_html=True)

elif buscar and not query.strip():
    st.warning("Escribe algo en el buscador para continuar.")

elif not query.strip():
    # Estado inicial: mostrar stats
    st.markdown('<div style="margin-top:16px"></div>', unsafe_allow_html=True)
    modelos_unicos = set()
    for m_list in df_rep['Modelos_lista']:
        modelos_unicos.update(m_list)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""<div class="metric-card blue">
            <div class="metric-value">{len(df_rep)}</div>
            <div class="metric-label">Repuestos en base</div></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value">{len(modelos_unicos)}</div>
            <div class="metric-label">Modelos cubiertos</div></div>""", unsafe_allow_html=True)
    with c3:
        fuentes = df_rep['Fuente'].nunique()
        st.markdown(f"""<div class="metric-card green">
            <div class="metric-value">{fuentes}</div>
            <div class="metric-label">Fuentes de datos</div></div>""", unsafe_allow_html=True)

    st.markdown('<div style="margin-top:28px"></div>', unsafe_allow_html=True)

    # Modelos disponibles
    st.markdown("#### Modelos con repuestos en la base")
    modelos_sorted = sorted(modelos_unicos)
    cols = st.columns(5)
    for i, modelo in enumerate(modelos_sorted):
        with cols[i % 5]:
            st.markdown(f"""
            <div style="background:#F0F4FF;color:#1A4FAE;padding:4px 10px;border-radius:8px;
                        font-size:0.82rem;font-weight:600;margin-bottom:6px;
                        cursor:pointer;text-align:center">{modelo}</div>
            """, unsafe_allow_html=True)

# ── DESCARGA BASE COMPLETA ────────────────────────────────────────────────────
st.markdown('<div style="margin-top:32px"></div>', unsafe_allow_html=True)
st.markdown('<hr style="border:none;border-top:1px solid #EEE;margin-bottom:20px">', unsafe_allow_html=True)

def generar_excel_base(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Base de Repuestos'

    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = 'XR MOTO STORE — BASE DE COMPATIBILIDADES DE REPUESTOS HONDA'
    c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='CC0000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers = ['Código', 'Descripción', 'Modelos Compatibles', 'Fuente']
    widths  = [20, 38, 80, 30]
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='404040')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    for ri, row in df.iterrows():
        er = ri + 3
        alt = ri % 2 == 0
        vals = [row['Codigo'], row['Descripcion'], row['Modelos'], row['Fuente']]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.border = border
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=(ci == 3))
            if alt:
                cell.fill = PatternFill('solid', start_color='F5F5F5')
            if ci == 1:
                cell.font = Font(name='Arial', bold=True, size=9, color='333333')
        ws.row_dimensions[er].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

col_dl, col_info = st.columns([1, 2])
with col_dl:
    clave = st.text_input("🔒 Clave de administrador", type="password",
                          placeholder="Ingresa la clave para descargar")
    if clave == "xr3010":
        excel_base = generar_excel_base(df_rep[['Codigo','Descripcion','Modelos','Fuente']])
        st.download_button(
            label="⬇️  Descargar base completa (Excel)",
            data=excel_base,
            file_name="XR_Base_Compatibilidades.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    elif clave:
        st.markdown('<p style="color:#CC0000;font-size:0.82rem;margin-top:4px">Clave incorrecta.</p>',
                    unsafe_allow_html=True)
with col_info:
    st.markdown(f"""
    <p style="color:#888;font-size:0.85rem;margin-top:8px">
        Descarga la base completa con <b>{len(df_rep)} repuestos</b>, sus códigos Honda,
        descripciones y todos los modelos compatibles encontrados hasta la fecha.
        Solo disponible para administradores.
    </p>
    """, unsafe_allow_html=True)

st.markdown('<div style="margin-top:16px"></div>', unsafe_allow_html=True)
with st.expander("ℹ️  ¿Cómo se amplía la base de datos?"):
    st.markdown("""
    La base actual tiene **302 repuestos** de la Lista Promo Mayo 2026.

    Se puede ampliar de dos formas:
    - **Automática**: cada vez que se carga una nueva lista promocional, los repuestos nuevos se incorporan.
    - **Manual**: el administrador agrega repuestos directamente editando el archivo `data/repuestos.csv` en GitHub.

    Con el tiempo, la base irá creciendo con cada lista y con repuestos frecuentes que no estén en promo.
    """)

render_footer()

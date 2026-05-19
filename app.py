import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
from datetime import datetime

# ── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="XR Pedidos Promo",
    page_icon="🏍️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── ESTILOS GLOBALES ─────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Fondo general */
.stApp { background-color: #1A1A1A; }

/* Header principal */
.xr-header {
    background: linear-gradient(135deg, #CC0000 0%, #8B0000 50%, #1A1A1A 100%);
    border-bottom: 3px solid #CC0000;
    padding: 18px 32px;
    border-radius: 12px;
    margin-bottom: 28px;
    display: flex;
    align-items: center;
    gap: 24px;
}
.xr-header h1 {
    color: #FFFFFF;
    font-size: 2rem;
    font-weight: 800;
    margin: 0;
    letter-spacing: -0.5px;
}
.xr-header p {
    color: #CCCCCC;
    font-size: 0.9rem;
    margin: 4px 0 0 0;
}

/* Tarjetas de carga */
.upload-card {
    background: #242424;
    border: 1.5px solid #333333;
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 12px;
    transition: border-color 0.2s;
}
.upload-card:hover { border-color: #CC0000; }
.upload-card h4 {
    color: #CC0000;
    font-size: 0.85rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin: 0 0 4px 0;
}
.upload-card p {
    color: #999999;
    font-size: 0.78rem;
    margin: 0 0 12px 0;
}

/* Métricas */
.metric-card {
    background: #242424;
    border-radius: 10px;
    padding: 18px 22px;
    border-left: 4px solid #CC0000;
    text-align: center;
}
.metric-card.blue  { border-left-color: #005BAC; }
.metric-card.green { border-left-color: #00A651; }
.metric-card.orange{ border-left-color: #FF8C00; }
.metric-value {
    font-size: 2rem;
    font-weight: 800;
    color: #FFFFFF;
    line-height: 1;
}
.metric-label {
    font-size: 0.75rem;
    color: #AAAAAA;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-top: 6px;
}

/* Botón principal */
.stButton > button {
    background: linear-gradient(135deg, #CC0000, #990000) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-size: 1.1rem !important;
    font-weight: 700 !important;
    padding: 14px 40px !important;
    width: 100% !important;
    letter-spacing: 0.5px !important;
    transition: all 0.2s !important;
    cursor: pointer !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #FF0000, #CC0000) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 20px rgba(204,0,0,0.4) !important;
}

/* Download button */
.stDownloadButton > button {
    background: linear-gradient(135deg, #005BAC, #003D7A) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-size: 1rem !important;
    font-weight: 700 !important;
    padding: 12px 32px !important;
    width: 100% !important;
}

/* Divider */
.xr-divider {
    border: none;
    border-top: 1px solid #333333;
    margin: 24px 0;
}

/* Badge estado */
.badge-sin-stock  { background:#C00000; color:white; padding:3px 10px; border-radius:12px; font-size:0.75rem; font-weight:700; }
.badge-bajo       { background:#9C5700; color:white; padding:3px 10px; border-radius:12px; font-size:0.75rem; font-weight:700; }
.badge-completar  { background:#7D6608; color:white; padding:3px 10px; border-radius:12px; font-size:0.75rem; font-weight:700; }
.badge-suficiente { background:#276221; color:white; padding:3px 10px; border-radius:12px; font-size:0.75rem; font-weight:700; }

/* Footer */
.xr-footer {
    text-align: center;
    color: #555555;
    font-size: 0.75rem;
    margin-top: 40px;
    padding-top: 16px;
    border-top: 1px solid #2A2A2A;
}
</style>
""", unsafe_allow_html=True)


# ── HEADER ───────────────────────────────────────────────────────────────────
logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
col_logo, col_title = st.columns([1, 5])
with col_logo:
    if os.path.exists(logo_path):
        st.image(logo_path, width=110)
with col_title:
    st.markdown("""
    <div style="padding-top:8px">
        <div style="font-size:1.9rem;font-weight:900;color:#FFFFFF;line-height:1">
            XR Pedidos Promo
        </div>
        <div style="font-size:0.9rem;color:#AAAAAA;margin-top:4px">
            Sistema de recomendación de compra — Repuestos Honda
        </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<div style="border-top:2px solid #CC0000;margin:16px 0 28px 0"></div>',
            unsafe_allow_html=True)


# ── ZONA DE CARGA ────────────────────────────────────────────────────────────
st.markdown("### 📂 Carga de archivos")
st.markdown('<p style="color:#999;font-size:0.85rem;margin-top:-12px">Sube los tres archivos para generar el análisis</p>',
            unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("""<div class="upload-card">
        <h4>📊 Historial de Ventas</h4>
        <p>DATA XR DESDE 2024 (.xls)</p></div>""", unsafe_allow_html=True)
    f_historial = st.file_uploader("", type=["xls","xlsx"], key="hist",
                                    label_visibility="collapsed")
    if f_historial:
        st.success(f"✅ {f_historial.name}")

with col2:
    st.markdown("""<div class="upload-card">
        <h4>🏷️ Lista Promocional</h4>
        <p>Lista Promo Honda (.xlsx)</p></div>""", unsafe_allow_html=True)
    f_promo = st.file_uploader("", type=["xlsx"], key="promo",
                                label_visibility="collapsed")
    if f_promo:
        st.success(f"✅ {f_promo.name}")

with col3:
    st.markdown("""<div class="upload-card">
        <h4>📦 Stock del Día</h4>
        <p>Stock actualizado (.xlsx)</p></div>""", unsafe_allow_html=True)
    f_stock = st.file_uploader("", type=["xlsx"], key="stock",
                                label_visibility="collapsed")
    if f_stock:
        st.success(f"✅ {f_stock.name}")

st.markdown('<hr class="xr-divider">', unsafe_allow_html=True)

# ── BOTÓN PRINCIPAL ───────────────────────────────────────────────────────────
archivos_ok = f_historial and f_promo and f_stock
if not archivos_ok:
    faltantes = []
    if not f_historial: faltantes.append("Historial")
    if not f_promo:     faltantes.append("Lista Promo")
    if not f_stock:     faltantes.append("Stock")
    st.info(f"⬆️  Sube los archivos faltantes para continuar: **{', '.join(faltantes)}**")

btn_col = st.columns([1, 2, 1])[1]
with btn_col:
    crear = st.button("🛒  CREAR PEDIDO", disabled=not archivos_ok)


# ── LÓGICA DE ANÁLISIS ────────────────────────────────────────────────────────
def analizar(f_hist, f_promo_, f_stk):

    # Cargar datos
    df_data = pd.read_excel(f_hist, sheet_name='vtadetart')
    df_data['fecha']  = pd.to_datetime(df_data['fecha'])
    df_data['codart'] = df_data['codart'].astype(str).str.strip()
    df_data['mes']    = df_data['fecha'].dt.to_period('M')

    df_stock_raw = pd.read_excel(f_stk, sheet_name='Hoja1', header=4)
    cod_col = 'CÓDIGO' if 'CÓDIGO' in df_stock_raw.columns else 'CODIGO'
    df_stock_raw = df_stock_raw[df_stock_raw[cod_col].notna()].copy()
    df_stock_raw[cod_col] = df_stock_raw[cod_col].astype(str).str.strip()
    df_stock_raw['STOCK'] = pd.to_numeric(df_stock_raw['STOCK'], errors='coerce').fillna(0)

    df_promo_raw = pd.read_excel(f_promo_, sheet_name='Data', header=5)
    df_promo_raw.columns = ['Codigo','Descripcion','Modelo','PrecioPublico','PrecioDealer',
                            'PrecioOfertaUSD','PrecioOfertaSoles','Desc','CantMinima',
                            'Pedido','PedidoUSD','col12','col13']
    df_promo_raw = df_promo_raw[
        df_promo_raw['Codigo'].notna() &
        (df_promo_raw['Codigo'].astype(str).str.strip() != 'NMaterial')].copy()
    df_promo_raw['Codigo']       = df_promo_raw['Codigo'].astype(str).str.strip()
    df_promo_raw['CantMinima']   = pd.to_numeric(df_promo_raw['CantMinima'],   errors='coerce').fillna(1)
    df_promo_raw['Desc']         = pd.to_numeric(df_promo_raw['Desc'],         errors='coerce').fillna(0)
    df_promo_raw['PrecioPublico']= pd.to_numeric(df_promo_raw['PrecioPublico'],errors='coerce').fillna(0)
    df_promo_raw['PrecioOfertaUSD'] = pd.to_numeric(df_promo_raw['PrecioOfertaUSD'], errors='coerce').fillna(0)

    # Meses completos (excluir mes actual incompleto)
    mes_actual       = df_data['mes'].max()
    months_completos = sorted([m for m in df_data['mes'].unique() if m < mes_actual])
    n_meses          = len(months_completos)

    monthly_sales   = df_data.groupby(['codart','mes'])['cantidad'].sum().reset_index()
    monthly_sales_c = monthly_sales[monthly_sales['mes'] < mes_actual]

    def get_ventas(code):
        all_m  = pd.DataFrame({'mes': months_completos})
        prod   = monthly_sales_c[monthly_sales_c['codart'] == code]
        return all_m.merge(prod, on='mes', how='left').fillna(0)['cantidad'].tolist()

    def limpiar_picos(vals):
        if len(vals) < 3:
            return vals, False
        vals_l, hubo = vals.copy(), False
        for i in range(len(vals)):
            otros = [v for j, v in enumerate(vals) if j != i]
            avg_o = np.mean(otros)
            if avg_o > 0 and vals[i] > avg_o * 3.0:
                vals_l[i] = avg_o
                hubo = True
        return vals_l, hubo

    def analizar_prod(code):
        vals_raw = get_ventas(code)
        if sum(vals_raw) == 0:
            return dict(trend='SIN HISTORIAL', avg=0, freq=0,
                        intermitente=False, pico=False)
        vals_l, pico = limpiar_picos(vals_raw)
        freq = sum(1 for v in vals_raw if v > 0)
        avg  = sum(vals_l) / n_meses
        n    = len(vals_l)
        if n >= 3:
            r, o = vals_l[-2:], vals_l[:-2]
            ar, ao = np.mean(r), np.mean(o) if o else 0
            if ao == 0 and ar > 0:           trend = 'CRECIENTE NUEVO'
            elif ao > 0:
                p = (ar - ao) / ao
                trend = 'CRECIENDO' if p>=0.20 else ('BAJANDO' if p<=-0.20 else 'ESTABLE')
            else: trend = 'ESTABLE'
        elif n == 2:
            trend = ('CRECIENDO' if vals_l[1]>vals_l[0]*1.20
                     else 'BAJANDO' if vals_l[1]<vals_l[0]*0.80 else 'ESTABLE')
        else: trend = 'ESTABLE'
        return dict(trend=trend, avg=round(avg,2), freq=freq,
                    intermitente=freq<=1, pico=pico)

    def get_stk(code):
        row = df_stock_raw[df_stock_raw[cod_col] == code]
        return float(row.iloc[0]['STOCK']) if len(row) else 0

    def cobertura(desc, interm):
        if interm:  return 2
        if desc >= 0.50: return 6
        if desc >= 0.40: return 4
        return 3

    def qty_rec(avg, stk, trend, interm, cant_min, desc):
        if avg == 0: return 0
        cob = cobertura(desc, interm)
        aj  = {'CRECIENDO':1.20,'CRECIENTE NUEVO':1.15,'ESTABLE':1.0,'BAJANDO':0.80}.get(trend,1.0)
        dem = avg * cob * aj
        q   = max(0.0, dem - stk)
        if q > 0:
            if cant_min <= q:           q = max(q, cant_min)
            elif cant_min <= q * 2.0:   q = cant_min
            elif avg >= cant_min / cob: q = cant_min
        return int(round(q))

    def estado(stk, q, avg):
        if stk == 0 and q > 0: return 'SIN STOCK / PEDIR'
        if stk > 0 and q > 0:
            return 'STOCK BAJO / PEDIR' if (avg > 0 and stk < avg) else 'COMPLETAR STOCK'
        return 'STOCK SUFICIENTE'

    def motivo(stk, q, trend, avg, freq, interm, pico, desc, cob):
        p = []
        p.append({'CRECIENDO':'Tendencia creciente','CRECIENTE NUEVO':'Tendencia nueva al alza',
                  'BAJANDO':'Tendencia a la baja - pedir con cautela',
                  'ESTABLE':'Rotacion estable'}.get(trend,''))
        if desc >= 0.50:  p.append('descuento promo alto (%.0f%%) - cobertura %dm' % (desc*100, cob))
        elif desc >= 0.40:p.append('descuento promo %.0f%% - cobertura %dm' % (desc*100, cob))
        else:             p.append('descuento %.0f%% - cobertura %dm' % (desc*100, cob))
        if interm: p.append('venta intermitente (%d/%d meses)' % (freq, n_meses))
        elif pico: p.append('pico aislado excluido del promedio')
        if avg > 0: p.append('promedio %.1f und/mes' % avg)
        if stk == 0: p.append('sin stock actual')
        elif q > 0:
            ca = round(stk/avg, 1) if avg > 0 else 0
            p.append('cobertura actual %.1f meses' % ca)
        else:
            ca = round(stk/avg, 1) if avg > 0 else 0
            p.append('stock cubre %.1f meses' % ca)
        if q > 0: p.append('recomendado pedir %d und' % q)
        return '; '.join(x for x in p if x).capitalize() + '.'

    def score_fn(avg, stk, trend, freq, interm, q, desc):
        s = avg * 10
        s *= {'CRECIENDO':1.6,'CRECIENTE NUEVO':1.3,'ESTABLE':1.0,
              'BAJANDO':0.55,'SIN HISTORIAL':0.0}.get(trend,1.0)
        if stk == 0 and trend != 'SIN HISTORIAL': s += 30
        elif stk > 0 and avg > 0 and stk < avg:   s += 15
        s += freq * 5 + desc * 60
        if interm: s *= 0.60
        if q == 0: s *= 0.20
        return round(s, 2)

    rows = []
    for _, pr in df_promo_raw.iterrows():
        code     = pr['Codigo']
        desc     = float(pr['Desc'])
        cant_min = float(pr['CantMinima'])
        info     = analizar_prod(code)
        stk      = get_stk(code)
        cob      = cobertura(desc, info['intermitente'])
        q        = qty_rec(info['avg'], stk, info['trend'], info['intermitente'], cant_min, desc)
        est      = estado(stk, q, info['avg'])
        mot      = motivo(stk, q, info['trend'], info['avg'], info['freq'],
                          info['intermitente'], info['pico'], desc, cob)
        sc       = score_fn(info['avg'], stk, info['trend'], info['freq'],
                            info['intermitente'], q, desc)
        rows.append({
            'Codigo': code, 'Descripcion': pr['Descripcion'],
            'Modelo aplicacion': pr['Modelo'], 'Tendencia': info['trend'],
            'Stock real': int(stk), 'Cant recomendada': q,
            'Estado compra': est, 'Motivo': mot,
            '_avg': info['avg'], '_desc': desc,
            '_precio_pub': float(pr['PrecioPublico']),
            '_precio_of': float(pr['PrecioOfertaUSD']),
            '_interm': info['intermitente'], '_score': sc,
            '_historial': info['trend'] != 'SIN HISTORIAL',
        })

    df_all = pd.DataFrame(rows)
    df_p   = df_all[df_all['Cant recomendada'] > 0].sort_values('_score', ascending=False)
    df_ok  = df_all[(df_all['Cant recomendada']==0) & df_all['_historial']].sort_values('_score', ascending=False)
    n_p    = min(len(df_p), 60)
    df_rep = pd.concat([df_p.head(n_p), df_ok.head(60-n_p)]).head(60).copy()

    ORDEN  = {'SIN STOCK / PEDIR':0,'STOCK BAJO / PEDIR':1,'COMPLETAR STOCK':2,'STOCK SUFICIENTE':3}
    df_rep['_orden'] = df_rep['Estado compra'].map(ORDEN)
    df_rep = df_rep.sort_values(['_orden','_score'], ascending=[True,False]).reset_index(drop=True)
    return df_rep


def generar_excel(df, fecha_str):
    COLOR_HDR = 'CC0000'
    COLOR_ALT = 'F5F5F5'
    ESTADO_COL = {
        'SIN STOCK / PEDIR':  ('FFC7CE','C00000'),
        'STOCK BAJO / PEDIR': ('FFEB9C','9C5700'),
        'COMPLETAR STOCK':    ('FFFF99','7D6608'),
        'STOCK SUFICIENTE':   ('C6EFCE','276221'),
    }
    TREND_COL = {
        'CRECIENDO':'276221','CRECIENTE NUEVO':'7030A0',
        'ESTABLE':'595959','BAJANDO':'C00000','SIN HISTORIAL':'808080',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RECOMENDACION COMPRA'

    # Titulo
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = 'XR MOTO STORE — TOP 60 RECOMENDACION DE COMPRA — LISTA PROMOCIONAL  |  %s' % fecha_str
    c.font  = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    c.fill  = PatternFill('solid', start_color=COLOR_HDR)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:J2')
    c = ws['A2']
    c.value = 'Cobertura: 6m (desc>50%) / 4m (40-50%) / 3m (<40%) / 2m si intermitente  |  Tendencia meses completos  |  Picos excluidos  |  Ajuste tendencia max +/-20%'
    c.font  = Font(name='Arial', italic=True, size=8, color='595959')
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 13

    headers    = ['Codigo','Descripcion','Modelo aplicacion','Tendencia','Stock real',
                  'Desc. Promo','Precio Oferta\n(USD)','Cantidad recomendada\na pedir',
                  'Estado compra','Motivo de Recomendacion']
    col_widths = [18,34,32,18,11,11,12,14,22,70]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color=COLOR_HDR)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 32

    thin   = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, row in df.iterrows():
        er     = ri + 4
        estado = row['Estado compra']
        bg, fc = ESTADO_COL.get(estado, ('FFFFFF','000000'))
        alt    = ri % 2 == 1
        desc   = row['_desc']
        if   desc >= 0.50: dbg, dfc = 'FFCCCC','C00000'
        elif desc >= 0.40: dbg, dfc = 'FFE5CC','9C5700'
        else:              dbg, dfc = 'FFFFCC','7D6608'

        vals = [row['Codigo'], row['Descripcion'], row['Modelo aplicacion'],
                row['Tendencia'], row['Stock real'],
                '%.0f%%' % (desc*100), row['_precio_of'],
                row['Cant recomendada'], estado, row['Motivo']]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.border = border
            if alt and ci not in [9]: cell.fill = PatternFill('solid', start_color=COLOR_ALT)
            if ci == 9:
                cell.fill = PatternFill('solid', start_color=bg)
                cell.font = Font(name='Arial', bold=True, size=9, color=fc)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif ci == 6:
                cell.fill = PatternFill('solid', start_color=dbg)
                cell.font = Font(name='Arial', bold=True, size=10, color=dfc)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 4:
                cell.font = Font(name='Arial', bold=True, size=9,
                                 color=TREND_COL.get(row['Tendencia'],'000000'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 8:
                color_c = 'C00000' if val > 0 else '276221'
                cell.font = Font(name='Arial', bold=True, size=11, color=color_c)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row['_interm'] and val > 0:
                    cell.fill = PatternFill('solid', start_color='FFF2CC')
            elif ci in [5, 7]:
                cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 1:
                cell.font = Font(name='Arial', bold=True, size=9)
            else:
                cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[er].height = 44 if len(str(row['Motivo'])) > 100 else 32

    # Hoja resumen
    ws2 = wb.create_sheet('RESUMEN')
    ws2.merge_cells('A1:E1')
    ws2['A1'].value = 'RESUMEN — XR MOTO STORE — TOP 60  |  %s' % fecha_str
    ws2['A1'].font  = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws2['A1'].fill  = PatternFill('solid', start_color=COLOR_HDR)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 24

    for ci, h in enumerate(['Estado','N productos','Unidades a pedir','Ahorro estimado (USD)','Desc. promedio'],1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='404040')
        c.alignment = Alignment(horizontal='center', vertical='center')

    for ri2, est in enumerate(['SIN STOCK / PEDIR','STOCK BAJO / PEDIR','COMPLETAR STOCK','STOCK SUFICIENTE'], 3):
        sub = df[df['Estado compra'] == est]
        bg2, fc2 = ESTADO_COL.get(est,('FFFFFF','000000'))
        ahorro = ((sub['_precio_pub'] - sub['_precio_of']) * sub['Cant recomendada'].clip(lower=0)).sum()
        dp = sub['_desc'].mean()*100 if len(sub) else 0
        for ci, val in enumerate([est, len(sub), int(sub['Cant recomendada'].sum()),
                                   round(ahorro,2), '%.1f%%' % dp], 1):
            c = ws2.cell(row=ri2, column=ci, value=val)
            c.fill = PatternFill('solid', start_color=bg2)
            c.font = Font(name='Arial', bold=(ci==1), size=10, color=fc2)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border

    sub_t  = df[df['Cant recomendada'] > 0]
    ah_t   = ((sub_t['_precio_pub'] - sub_t['_precio_of']) * sub_t['Cant recomendada']).sum()
    for ci, val in enumerate(['TOTAL', len(sub_t), int(sub_t['Cant recomendada'].sum()),
                               round(ah_t,2), '%.1f%%' % (df['_desc'].mean()*100)],1):
        c = ws2.cell(row=7, column=ci, value=val)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1F3864')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    for ci2, w in enumerate([28,14,18,20,14],1):
        ws2.column_dimensions[get_column_letter(ci2)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── EJECUCIÓN ─────────────────────────────────────────────────────────────────
if crear and archivos_ok:
    st.markdown('<hr class="xr-divider">', unsafe_allow_html=True)

    with st.spinner(""):
        prog = st.progress(0)
        status = st.empty()

        status.markdown("⏳ **Cargando historial de ventas...**")
        prog.progress(20)

        status.markdown("⏳ **Cruzando con stock actual...**")
        prog.progress(45)

        status.markdown("⏳ **Analizando tendencias y descuentos...**")
        prog.progress(70)

        df_resultado = analizar(f_historial, f_promo, f_stock)

        status.markdown("⏳ **Generando Excel...**")
        prog.progress(90)

        fecha_str = datetime.now().strftime("%d/%m/%Y")
        excel_buf = generar_excel(df_resultado, fecha_str)

        prog.progress(100)
        status.markdown("✅ **¡TOP 60 generado exitosamente!**")

    st.markdown('<hr class="xr-divider">', unsafe_allow_html=True)

    # ── MÉTRICAS ──────────────────────────────────────────────────────────────
    st.markdown("### 📊 Resumen del análisis")
    df_pedir = df_resultado[df_resultado['Cant recomendada'] > 0]

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value">{len(df_resultado)}</div>
            <div class="metric-label">Productos analizados</div></div>""",
            unsafe_allow_html=True)
    with m2:
        n_sin = len(df_resultado[df_resultado['Estado compra']=='SIN STOCK / PEDIR'])
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value" style="color:#FF4444">{n_sin}</div>
            <div class="metric-label">Sin stock — pedir urgente</div></div>""",
            unsafe_allow_html=True)
    with m3:
        total_u = int(df_pedir['Cant recomendada'].sum())
        st.markdown(f"""<div class="metric-card blue">
            <div class="metric-value">{total_u:,}</div>
            <div class="metric-label">Unidades a pedir</div></div>""",
            unsafe_allow_html=True)
    with m4:
        ahorro = ((df_pedir['_precio_pub'] - df_pedir['_precio_of']) * df_pedir['Cant recomendada']).sum()
        st.markdown(f"""<div class="metric-card green">
            <div class="metric-value">$ {ahorro:,.0f}</div>
            <div class="metric-label">Ahorro estimado (USD)</div></div>""",
            unsafe_allow_html=True)

    st.markdown('<div style="margin-top:24px"></div>', unsafe_allow_html=True)

    # ── TABLA INTERACTIVA ────────────────────────────────────────────────────
    st.markdown("### 📋 TOP 60 — Vista previa")

    filtro = st.multiselect(
        "Filtrar por estado:",
        options=['SIN STOCK / PEDIR','STOCK BAJO / PEDIR','COMPLETAR STOCK','STOCK SUFICIENTE'],
        default=['SIN STOCK / PEDIR','STOCK BAJO / PEDIR','COMPLETAR STOCK'],
    )

    df_vista = df_resultado[df_resultado['Estado compra'].isin(filtro)] if filtro else df_resultado

    COLORES_ESTADO = {
        'SIN STOCK / PEDIR':  'background-color: #FFC7CE; color: #C00000',
        'STOCK BAJO / PEDIR': 'background-color: #FFEB9C; color: #9C5700',
        'COMPLETAR STOCK':    'background-color: #FFFF99; color: #7D6608',
        'STOCK SUFICIENTE':   'background-color: #C6EFCE; color: #276221',
    }

    def color_estado(val):
        return COLORES_ESTADO.get(val, '')

    df_show = df_vista[['Codigo','Descripcion','Tendencia','Stock real',
                         'Cant recomendada','Estado compra']].copy()
    df_show.columns = ['Código','Descripción','Tendencia','Stock','Cant. Pedir','Estado']
    df_show = df_show.reset_index(drop=True)
    df_show.index += 1

    st.dataframe(
        df_show.style.map(color_estado, subset=['Estado']),
        use_container_width=True,
        height=420,
    )

    st.markdown('<div style="margin-top:20px"></div>', unsafe_allow_html=True)

    # ── DESCARGA ─────────────────────────────────────────────────────────────
    nombre_archivo = 'XR_Pedido_Promo_%s.xlsx' % datetime.now().strftime('%Y%m%d')
    dl_col = st.columns([1, 2, 1])[1]
    with dl_col:
        st.download_button(
            label="⬇️  DESCARGAR EXCEL",
            data=excel_buf,
            file_name=nombre_archivo,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

# ── FOOTER ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="xr-footer">
    XR Moto Store S.A.C. — Sistema de Recomendación de Compras — Repuestos Honda
</div>
""", unsafe_allow_html=True)
